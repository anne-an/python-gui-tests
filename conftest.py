import os
import openpyxl
import pytest
from fixture.application import Application


@pytest.fixture(scope="session")
def app(request):
    fixture = Application("C:\\Users\\anne_\\Documents\\AddressBook\\AddressBook.exe")
    request.addfinalizer(fixture.destroy)
    return fixture

def pytest_generate_tests(metafunc):
    for fixture in metafunc.fixturenames:
        if fixture.startswith("xl_"):
            print("FFF " + fixture)
            testdata = load_from_excel(fixture[3:])
            metafunc.parametrize(fixture, testdata, ids=[str(x) for x in testdata])

def load_from_excel(file):
    list = []
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), "generator\%s.xlsx" % file))
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        for cell in range(1, ws.max_column + 1):
            list.append(ws.cell(row, cell))
    return list
